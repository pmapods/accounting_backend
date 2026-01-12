# generate_inventory_report.py - WITH BASO SUPPORT
# Modified: Added BASO file processing with 4 sheets (GT GS, GT BS, MT GS, MT BS)
# BASO mapping: (Plant, Kode Barang) -> FISIK (PCS)
# Output: 3 new columns CC (GS BASO), CD (BS BASO), CE (Grand Total)

import sys
import json
import os
import datetime
import traceback
from collections import defaultdict
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from concurrent.futures import ThreadPoolExecutor
import warnings

def log(msg):
    """Log to stderr"""
    print(f"[worker] {msg}", file=sys.stderr, flush=True)

def find_col(df_cols, candidates):
    """Find first matching column name"""
    lower_map = {str(c).strip().lower(): c for c in df_cols if str(c).strip() and str(c).strip() != 'nan'}
    
    for cand in candidates:
        cand_lower = str(cand).lower()
        if cand_lower in lower_map:
            return lower_map[cand_lower]
    
    for cand in candidates:
        cand_lower = str(cand).lower()
        for key, original in lower_map.items():
            if cand_lower in key:
                return original
    return None

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def get_column_index(letter):
    """Convert Excel column letter to index (1-based)"""
    if len(letter) == 1:
        return ord(letter) - ord('A') + 1
    return sum([(ord(ch) - 64) * (26 ** (len(letter)-i-1)) for i, ch in enumerate(letter)])

def read_sheet(file_path, sheet_name):
    """Read single sheet - parallelizable"""
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
            df.columns = [str(c).strip() if not pd.isna(c) else f"Unnamed_{i}" 
                         for i, c in enumerate(df.columns)]
            return sheet_name, df
    except Exception as e:
        log(f"Error reading {sheet_name}: {str(e)}")
        return sheet_name, None

class SheetCache:
    """Cache for sheet lookups - build once, query many times"""
    
    def __init__(self, sheets_dict, baso_path=None):
        self.sheets = sheets_dict
        self.caches = {}
        self.baso_path = baso_path
        self._build_caches()
        if baso_path:
            self._build_baso_cache()
    
    def _build_caches(self):
        """Pre-build all lookup caches"""
        log("Building sheet lookup caches...")
        
        # Cache SALDO AWAL
        if 'SALDO AWAL' in self.sheets:
            df = self.sheets['SALDO AWAL']
            cols = list(df.columns)
            mat_col = find_col(cols, ["Kode Material", "Material"])
            plant_col = find_col(cols, ["Plant", "Plnt"])
            sloc_col = find_col(cols, ["Storage Loc", "Storage Location"])
            amt_col = find_col(cols, ["Closing Stock (pcs)", "QTY", "Closing Stock"])
            
            if mat_col and amt_col:
                df_clean = df.copy()
                df_clean['material'] = df_clean[mat_col].astype(str).str.strip()
                df_clean['plant'] = df_clean[plant_col].astype(str).str.strip() if plant_col else ''
                df_clean['sloc'] = df_clean[sloc_col].astype(str).str.strip() if sloc_col else ''
                df_clean['amount'] = pd.to_numeric(df_clean[amt_col], errors='coerce').fillna(0)
                
                grouped = df_clean.groupby(['material', 'plant', 'sloc'], dropna=False)['amount'].sum()
                self.caches['saldo_awal'] = grouped.to_dict()
                log(f"  SALDO AWAL cache: {len(self.caches['saldo_awal'])} entries")
        
        # Cache SALDO AWAL MB5B
        if 'SALDO AWAL MB5B' in self.sheets:
            df = self.sheets['SALDO AWAL MB5B']
            
            header_row_idx = None
            if len(df) > 1:
                for i in range(min(10, len(df))):
                    row_values = df.iloc[i].astype(str).str.lower().tolist()
                    if 'material' in ' '.join(row_values):
                        header_row_idx = i
                        break
            
            if header_row_idx is not None:
                df.columns = df.iloc[header_row_idx]
                df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
            
            cols = list(df.columns)
            mat_col = find_col(cols, ["Material"])
            plant_col = find_col(cols, ["Plnt", "Plant"])
            gs_col = find_col(cols, ["GS"])
            bs_col = find_col(cols, ["BS"])
            
            if mat_col and (gs_col or bs_col):
                df_clean = df.copy()
                df_clean['material'] = df_clean[mat_col].astype(str).str.strip()
                df_clean['plant'] = df_clean[plant_col].astype(str).str.strip() if plant_col else ''
                
                if gs_col:
                    df_clean['gs_amount'] = pd.to_numeric(df_clean[gs_col], errors='coerce').fillna(0)
                    grouped_gs = df_clean.groupby(['material', 'plant'], dropna=False)['gs_amount'].sum()
                    self.caches['mb5b_awal_gs'] = grouped_gs.to_dict()
                    log(f"  MB5B AWAL GS cache: {len(self.caches['mb5b_awal_gs'])} entries")
                
                if bs_col:
                    df_clean['bs_amount'] = pd.to_numeric(df_clean[bs_col], errors='coerce').fillna(0)
                    grouped_bs = df_clean.groupby(['material', 'plant'], dropna=False)['bs_amount'].sum()
                    self.caches['mb5b_awal_bs'] = grouped_bs.to_dict()
                    log(f"  MB5B AWAL BS cache: {len(self.caches['mb5b_awal_bs'])} entries")
        
        # Cache MB5B
        if '13. MB5B' in self.sheets:
            df = self.sheets['13. MB5B']
            cols = list(df.columns)
            mat_col = find_col(cols, ["Material"])
            plant_col = find_col(cols, ["Plnt", "Plant"])
            gs_col = find_col(cols, ["GS"])
            bs_col = find_col(cols, ["BS"])
            
            if mat_col and (gs_col or bs_col):
                df_clean = df.copy()
                df_clean['material'] = df_clean[mat_col].astype(str).str.strip()
                df_clean['plant'] = df_clean[plant_col].astype(str).str.strip() if plant_col else ''
                
                if gs_col:
                    df_clean['gs_amount'] = pd.to_numeric(df_clean[gs_col], errors='coerce').fillna(0)
                    grouped_gs = df_clean.groupby(['material', 'plant'], dropna=False)['gs_amount'].sum()
                    self.caches['mb5b_gs'] = grouped_gs.to_dict()
                    log(f"  MB5B GS cache: {len(self.caches['mb5b_gs'])} entries")
                
                if bs_col:
                    df_clean['bs_amount'] = pd.to_numeric(df_clean[bs_col], errors='coerce').fillna(0)
                    grouped_bs = df_clean.groupby(['material', 'plant'], dropna=False)['bs_amount'].sum()
                    self.caches['mb5b_bs'] = grouped_bs.to_dict()
                    log(f"  MB5B BS cache: {len(self.caches['mb5b_bs'])} entries")
        
        # Cache EDS
        if '14. SALDO AKHIR EDS' in self.sheets:
            df = self.sheets['14. SALDO AKHIR EDS']
            log(f"  === DEBUGGING EDS SHEET ===")
            log(f"  Sheet shape: {df.shape} (rows x cols)")
            
            cols = list(df.columns)
            mat_col = find_col(cols, ["Material"])
            plant_col = find_col(cols, ["Plant"])
            sloc_col = find_col(cols, ["Storage Location", "Storage Loc"])
            amt_col = find_col(cols, ["Closing Stock (pcs)", "Closing Stock", "QTY"])
            
            if mat_col and amt_col:
                df_clean = df.copy()
                df_clean['material'] = df_clean[mat_col].astype(str).str.strip()
                df_clean['plant'] = df_clean[plant_col].astype(str).str.strip() if plant_col else ''
                df_clean['sloc'] = df_clean[sloc_col].astype(str).str.strip() if sloc_col else ''
                df_clean['amount'] = pd.to_numeric(df_clean[amt_col], errors='coerce').fillna(0)
                
                grouped = df_clean.groupby(['material', 'plant', 'sloc'], dropna=False)['amount'].sum()
                self.caches['eds'] = grouped.to_dict()
                log(f"  EDS cache: {len(self.caches['eds'])} entries")

    def _build_baso_cache(self):
        """Build BASO cache from 4 sheets"""
        log("=== Building BASO cache ===")
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.baso_path, read_only=True, data_only=True)
            available_sheets = wb.sheetnames
            log(f"  Available BASO sheets: {available_sheets}")
            
            # Initialize caches
            self.caches['baso_gs'] = {}
            self.caches['baso_bs'] = {}
            
            # Process each sheet
            for sheet_name in available_sheets:
                # Detect type by sheet name ending
                sheet_lower = sheet_name.lower()
                if sheet_lower.endswith('gs'):
                    target_type = 'GS'
                elif sheet_lower.endswith('bs'):
                    target_type = 'BS'
                else:
                    log(f"  Skipping sheet '{sheet_name}' (not ending with GS or BS)")
                    continue
                
                log(f"  Processing BASO sheet: '{sheet_name}' -> {target_type}")
                
                # Read sheet
                df = pd.read_excel(self.baso_path, sheet_name=sheet_name, dtype=str)
                df.columns = [str(c).strip() if not pd.isna(c) else f"Unnamed_{i}" 
                             for i, c in enumerate(df.columns)]
                
                log(f"    Sheet shape: {df.shape}")
                log(f"    First 5 rows preview:")
                for i in range(min(5, len(df))):
                    log(f"      Row {i}: {list(df.iloc[i].values)[:8]}")
                
                # Find header row (row 3 = index 2)
                header_row_idx = None
                for i in range(min(10, len(df))):
                    row_values = df.iloc[i].astype(str).str.lower().tolist()
                    row_str = ' '.join(row_values)
                    if 'kode barang' in row_str or 'fisik' in row_str:
                        header_row_idx = i
                        log(f"    Found header at row {i}")
                        break
                
                if header_row_idx is None:
                    log(f"    ERROR: Could not find header row")
                    continue
                
                # Set header and remove rows before it
                df.columns = df.iloc[header_row_idx]
                df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
                
                # Find columns
                cols = list(df.columns)
                log(f"    Columns after header: {cols}")
                
                plant_col = find_col(cols, ["PLANT", "Plant"])
                material_col = find_col(cols, ["KODE BARANG", "Kode Barang", "Material"])
                fisik_col = find_col(cols, ["FISIK (PCS)", "Fisik (pcs)", "FISIK"])
                
                log(f"    Column mapping:")
                log(f"      Plant: {plant_col}")
                log(f"      Material: {material_col}")
                log(f"      Fisik: {fisik_col}")
                
                if not plant_col or not material_col or not fisik_col:
                    log(f"    ERROR: Missing required columns")
                    continue
                
                # Process data
                df_clean = df.copy()
                df_clean['plant'] = df_clean[plant_col].astype(str).str.strip().str.upper()
                df_clean['material'] = df_clean[material_col].astype(str).str.strip()
                df_clean['fisik'] = pd.to_numeric(df_clean[fisik_col], errors='coerce').fillna(0)
                
                # Remove empty rows
                df_clean = df_clean[
                    (df_clean['plant'] != '') & 
                    (df_clean['plant'] != 'NAN') &
                    (df_clean['material'] != '') & 
                    (df_clean['material'] != 'NAN')
                ]
                
                log(f"    Valid rows: {len(df_clean)}")
                
                # Group by (plant, material) and sum
                grouped = df_clean.groupby(['plant', 'material'], dropna=False)['fisik'].sum()
                
                # Store in appropriate cache
                cache_key = 'baso_gs' if target_type == 'GS' else 'baso_bs'
                for (plant, material), fisik_value in grouped.items():
                    key = (material, plant)
                    if key in self.caches[cache_key]:
                        self.caches[cache_key][key] += fisik_value
                    else:
                        self.caches[cache_key][key] = fisik_value
                
                log(f"    Added {len(grouped)} entries to {cache_key}")
            
            wb.close()
            
            log(f"  BASO GS cache: {len(self.caches.get('baso_gs', {}))} total entries")
            log(f"  BASO BS cache: {len(self.caches.get('baso_bs', {}))} total entries")
            
            # Show sample data
            if len(self.caches.get('baso_gs', {})) > 0:
                log(f"  Sample BASO GS entries (first 5):")
                for i, (key, val) in enumerate(list(self.caches['baso_gs'].items())[:5]):
                    log(f"    {key} = {val}")
            
            if len(self.caches.get('baso_bs', {})) > 0:
                log(f"  Sample BASO BS entries (first 5):")
                for i, (key, val) in enumerate(list(self.caches['baso_bs'].items())[:5]):
                    log(f"    {key} = {val}")
                    
        except Exception as e:
            log(f"  ERROR building BASO cache: {str(e)}")
            log(f"  Traceback: {traceback.format_exc()}")
            self.caches['baso_gs'] = {}
            self.caches['baso_bs'] = {}

    def get_saldo_awal(self, material, plant, sloc_type):
        if 'saldo_awal' not in self.caches:
            return 0.0
        return self.caches['saldo_awal'].get((material, plant, sloc_type), 0.0)
    
    def get_mb5b_awal(self, material, plant, sloc_type):
        cache_key = 'mb5b_awal_gs' if sloc_type == "GS" else 'mb5b_awal_bs'
        if cache_key not in self.caches:
            return 0.0
        return self.caches[cache_key].get((material, plant), 0.0)
    
    def get_mb5b(self, material, plant, sloc_type):
        cache_key = 'mb5b_gs' if sloc_type == "GS" else 'mb5b_bs'
        if cache_key not in self.caches:
            return 0.0
        return self.caches[cache_key].get((material, plant), 0.0)
    
    def get_eds(self, material, plant, sloc_type):
        if 'eds' not in self.caches:
            return 0.0
        return self.caches['eds'].get((material, plant, sloc_type), 0.0)
    
    def get_baso(self, material, plant, sloc_type):
        """Get BASO value for material+plant"""
        cache_key = 'baso_gs' if sloc_type == "GS" else 'baso_bs'
        if cache_key not in self.caches:
            return 0.0
        return self.caches[cache_key].get((material, plant), 0.0)

def main():
    try:
        payload = json.load(sys.stdin)
        files = payload.get("files", {})
        master_inventory = payload.get("master_inventory", [])
        master_movement = payload.get("master_movement", [])
        report_date = payload.get("report_date")

        mb51_path = files.get("mb51")
        main_path = files.get("main")
        baso_path = files.get("baso")  # TAMBAHAN: Path BASO (opsional)

        if not mb51_path or not main_path:
            raise ValueError("Payload must include files.mb51 and files.main paths")
        
        if not report_date:
            raise ValueError("Payload must include report_date")

        # Log BASO status
        if baso_path:
            log(f"BASO file provided: {baso_path}")
        else:
            log("BASO file not provided - will use zeros for BASO columns")

        # Load master data
        log("Loading master data...")
        df_master_inv = pd.DataFrame(master_inventory)
        df_master_mov = pd.DataFrame(master_movement)

        df_master_inv.columns = [str(c).strip().lower() if not pd.isna(c) else f"col_{i}" 
                                  for i, c in enumerate(df_master_inv.columns)]
        df_master_mov.columns = [str(c).strip().lower() if not pd.isna(c) else f"col_{i}" 
                                  for i, c in enumerate(df_master_mov.columns)]

        # Inventory mapping
        log("Creating inventory lookup...")
        df_master_inv['plant'] = df_master_inv['plant'].astype(str).str.strip()
        inv_map = df_master_inv.set_index('plant')[['area', 'kode_dist', 'profit_center']].to_dict('index')
        
        # Build mv_text -> mv_grouping mapping
        log("Building movement text -> grouping mapping...")
        df_master_mov['mv_text'] = df_master_mov['mv_text'].astype(str).str.strip().str.lower()
        df_master_mov['mv_grouping'] = df_master_mov['mv_grouping'].astype(str).str.strip()
        
        mv_text_to_grouping = {}
        for _, row in df_master_mov.iterrows():
            mv_text = row['mv_text']
            mv_grouping = row['mv_grouping']
            if mv_text and mv_text != 'nan' and mv_grouping and mv_grouping != 'nan':
                if mv_text not in mv_text_to_grouping:
                    mv_text_to_grouping[mv_text] = mv_grouping
        
        log(f"  Created {len(mv_text_to_grouping)} mv_text -> mv_grouping mappings")
        
        # Build storage + mv_grouping -> column mapping
        log("Building (storage, mv_grouping) -> column mapping...")
        storage_grouping_to_column = {
            # GS00 - 9 columns (R to Z)
            ("GS00", "Terima Barang"): "R",
            ("GS00", "Retur Beli"): "S",
            ("GS00", "Penjualan"): "T",
            ("GS00", "Retur Jual"): "U",
            ("GS00", "Intra Gudang Masuk"): "V",
            ("GS00", "Intra Gudang"): "W",
            ("GS00", "Transfer Stock"): "X",
            ("GS00", "Pemusnahan"): "Y",
            ("GS00", "Adjustment"): "Z",
            
            # BS00 - 9 columns (AB to AJ)
            ("BS00", "Terima Barang"): "AB",
            ("BS00", "Retur Beli"): "AC",
            ("BS00", "Penjualan"): "AD",
            ("BS00", "Retur Jual"): "AE",
            ("BS00", "Intra Gudang Masuk"): "AF",
            ("BS00", "Intra Gudang"): "AG",
            ("BS00", "Transfer Stock"): "AH",
            ("BS00", "Pemusnahan"): "AI",
            ("BS00", "Adjustment"): "AJ",
            
            # AI00 - 9 columns (AL to AT)
            ("AI00", "Terima Barang"): "AL",
            ("AI00", "Retur Beli"): "AM",
            ("AI00", "Penjualan"): "AN",
            ("AI00", "Retur Jual"): "AO",
            ("AI00", "Intra Gudang Masuk"): "AP",
            ("AI00", "Intra Gudang"): "AQ",
            ("AI00", "Transfer Stock"): "AR",
            ("AI00", "Pemusnahan"): "AS",
            ("AI00", "Adjustment"): "AT",
            
            # TR00 - 9 columns (AV to BD)
            ("TR00", "Terima Barang"): "AV",
            ("TR00", "Retur Beli"): "AW",
            ("TR00", "Penjualan"): "AX",
            ("TR00", "Retur Jual"): "AY",
            ("TR00", "Intra Gudang Masuk"): "AZ",
            ("TR00", "Intra Gudang"): "BA",
            ("TR00", "Transfer Stock"): "BB",
            ("TR00", "Pemusnahan"): "BC",
            ("TR00", "Adjustment"): "BD",
            
            # 641/642 (empty storage) - 2 columns (BF, BG)
            ("EMPTY_STORAGE", "Intra Gudang"): "BF",
            ("EMPTY_STORAGE", "Intra Gudang"): "BG",
        }
        
        log(f"  Created {len(storage_grouping_to_column)} (storage, mv_grouping) -> column mappings")

        # Read MB51
        log(f"Reading MB51...")
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            df_mb51 = pd.read_excel(mb51_path, sheet_name=0, engine="openpyxl", dtype=str)
        
        df_mb51.columns = [str(c).strip() if not pd.isna(c) else f"Unnamed_{i}" 
                          for i, c in enumerate(df_mb51.columns)]

        # Find columns
        mb_cols = list(df_mb51.columns)
        col_posting = find_col(mb_cols, ["Posting Date"])
        col_material = find_col(mb_cols, ["Material"])
        col_plant = find_col(mb_cols, ["Plant", "Plnt"])
        col_movement = find_col(mb_cols, ["Movement type", "Movement Type"])
        col_movement_text = find_col(mb_cols, ["Movement Type Text"])
        col_amount = find_col(mb_cols, ["Quantity"])
        col_sloc = find_col(mb_cols, ["Storage", "Storage Location", "Storage Loc"])
        col_material_desc = find_col(mb_cols, ["Material description"])

        missing = []
        for name, col in [("Posting Date", col_posting), ("Material", col_material), 
                         ("Plant", col_plant), ("Movement type", col_movement),
                         ("Movement Type Text", col_movement_text), ("Quantity", col_amount)]:
            if not col:
                missing.append(name)
        
        if missing:
            raise ValueError(f"Missing MB51 columns: {', '.join(missing)}")

        # Rename
        rename_dict = {
            col_posting: "posting_date", col_material: "material",
            col_plant: "plant", col_movement: "mv_type",
            col_movement_text: "mv_text", col_amount: "amount"
        }
        if col_sloc:
            rename_dict[col_sloc] = "sloc"
        if col_material_desc:
            rename_dict[col_material_desc] = "material_desc_mb51"
        
        df_mb51 = df_mb51.rename(columns=rename_dict)

        # Determine report period
        log("Determining report period from request body...")
        report_month_dt = datetime.datetime.strptime(report_date, "%Y-%m-%d")
        
        bulan = report_month_dt.strftime("%B").upper()
        tahun = report_month_dt.year
        prev_month_dt = report_month_dt
        prev_month = prev_month_dt.strftime("%B").upper()
        prev_year = prev_month_dt.year
        bulan_only = bulan

        log(f"  Report period from request: {bulan} {tahun}")

        # Convert data types
        log("Converting data types...")
        
        try:
            date_numeric = pd.to_numeric(df_mb51["posting_date"], errors='coerce')
            df_mb51["posting_date"] = pd.to_datetime(
                date_numeric,
                origin='1899-12-30',
                unit='D',
                errors='coerce'
            )
            
            valid_dates = df_mb51["posting_date"].notna().sum()
            log(f"  ✓ Converted {valid_dates}/{len(df_mb51)} dates successfully")
                
        except Exception as e:
            log(f"  ERROR converting dates: {str(e)}")
            df_mb51["posting_date"] = pd.NaT
        
        df_mb51["amount"] = pd.to_numeric(df_mb51["amount"], errors="coerce")
        
        neg_count = (df_mb51["amount"] < 0).sum()
        pos_count = (df_mb51["amount"] > 0).sum()
        total_amount = df_mb51["amount"].sum()
        
        log(f"  === AMOUNT DISTRIBUTION (RAW MB51) ===")
        log(f"  Positive: {pos_count}, Negative: {neg_count}")
        log(f"  Total sum: {total_amount:,.2f}")
        
        df_mb51["plant"] = df_mb51["plant"].astype(str).str.strip()
        df_mb51["mv_type"] = df_mb51["mv_type"].astype(str).str.strip()
        df_mb51["material"] = df_mb51["material"].astype(str).str.strip()
        df_mb51["mv_text"] = df_mb51["mv_text"].astype(str).str.strip().str.lower()
        
        # Handle storage
        if 'sloc' in df_mb51.columns:
            df_mb51["is_empty_storage"] = (
                (df_mb51["sloc"].isna()) |
                (df_mb51["sloc"].isnull()) |
                (df_mb51["sloc"].astype(str).str.strip() == '') |
                (df_mb51["sloc"].astype(str).str.strip().str.upper() == 'NAN') |
                (df_mb51["sloc"].astype(str).str.strip().str.upper() == 'NONE')
            )
            
            df_mb51["storage"] = df_mb51["sloc"].astype(str).str.strip().str.upper()
            df_mb51.loc[df_mb51["is_empty_storage"], "storage"] = 'EMPTY_STORAGE'
        else:
            df_mb51["storage"] = "EMPTY_STORAGE"
            df_mb51["is_empty_storage"] = True

        # Map inventory
        log("Mapping inventory data...")
        df_inv_lookup = pd.DataFrame([
            {'plant': k, **v} for k, v in inv_map.items()
        ])
        df_mb51 = df_mb51.merge(
            df_inv_lookup[['plant', 'area', 'kode_dist', 'profit_center']], 
            on='plant', how='left'
        )

        # Read main file sheets
        log("Reading main file sheets...")
        required_sheets = ['SALDO AWAL', 'SALDO AWAL MB5B', '13. MB5B', 
                          '14. SALDO AKHIR EDS', 'Output Report INV ARUS BARANG']
        
        sheets_dict = {}
        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = [executor.submit(read_sheet, main_path, sheet) for sheet in required_sheets]
            for future in futures:
                sheet_name, df = future.result()
                if df is not None:
                    sheets_dict[sheet_name] = df
                    log(f"  ✓ Loaded '{sheet_name}': {df.shape}")

        # Initialize sheet cache WITH BASO
        sheet_cache = SheetCache(sheets_dict, baso_path)

        # Get existing materials
        existing_materials = []
        if 'Output Report INV ARUS BARANG' in sheets_dict:
            df_existing = sheets_dict['Output Report INV ARUS BARANG']
            if df_existing.shape[0] > 8 and df_existing.shape[1] >= 7:
                df_existing_materials = df_existing.iloc[7:, [1, 5]].copy()
                df_existing_materials.columns = ['plant', 'material']
                df_existing_materials = df_existing_materials[
                    (df_existing_materials['material'].notna()) & 
                    (df_existing_materials['material'].astype(str).str.strip() != '') &
                    (df_existing_materials['material'].astype(str).str.strip() != 'nan')
                ]
                
                log("Loading existing materials from main file...")
                for idx, row in df_existing_materials.iterrows():
                    plant = str(row['plant']).strip().upper()
                    material = str(row['material']).strip()
                    
                    if plant in inv_map:
                        existing_materials.append({
                            'material': material, 'plant': plant,
                            'area': inv_map[plant]['area'],
                            'kode_dist': inv_map[plant]['kode_dist'],
                            'profit_center': inv_map[plant]['profit_center']
                        })
                    else:
                        existing_materials.append({
                            'material': material, 'plant': plant,
                            'area': '', 'kode_dist': '', 'profit_center': ''
                        })
                
                log(f"  Found {len(existing_materials)} existing materials")

        # Map unknown storages to GS00
        df_mb51_filtered = df_mb51.copy()
        log(f"MB51 data before storage mapping: {len(df_mb51_filtered)} rows")
        
        known_storages = ['GS00', 'BS00', 'AI00', 'TR00', 'EMPTY_STORAGE']
        unknown_storage_mask = ~df_mb51_filtered['storage'].isin(known_storages)
        unknown_storage_count = unknown_storage_mask.sum()
        
        if unknown_storage_count > 0:
            log(f"Found {unknown_storage_count} rows with unknown storage")
            unknown_storages = df_mb51_filtered[unknown_storage_mask]['storage'].unique()
            log(f"  Unknown storages: {list(unknown_storages)[:10]}")
            df_mb51_filtered.loc[unknown_storage_mask, 'storage'] = 'GS00'
            log(f"  → Mapped all unknown storages to GS00")
        
        log(f"After storage mapping: {len(df_mb51_filtered)} rows")
        
        df_mb51_filtered['plant_clean'] = df_mb51_filtered['plant'].astype(str).str.strip().str.upper()
        mb51_plants = set(df_mb51_filtered['plant_clean'].unique())
        log(f"Unique plants in MB51: {len(mb51_plants)}")

        # Group MB51
        log("=== Grouping MB51 by exact combination ===")
        grouped_mb51 = df_mb51_filtered.groupby(
            ['material', 'plant_clean', 'storage', 'mv_type', 'mv_text'],
            dropna=False
        ).agg({'amount': 'sum'}).reset_index()
        
        log(f"  Grouped MB51: {len(grouped_mb51)} unique combinations")
        
        # Map mv_text to mv_grouping
        log("Mapping mv_text to mv_grouping...")
        grouped_mb51['mv_grouping'] = grouped_mb51['mv_text'].map(mv_text_to_grouping)
        
        mapped_count = grouped_mb51['mv_grouping'].notna().sum()
        unmapped_count = grouped_mb51['mv_grouping'].isna().sum()
        log(f"  Mapped: {mapped_count}/{len(grouped_mb51)}")
        log(f"  Unmapped: {unmapped_count}/{len(grouped_mb51)}")
        
        # Determine target column
        log("Determining target columns...")
        
        def get_target_column(row):
            storage = row['storage']
            mv_grouping = row['mv_grouping']
            mv_type = row['mv_type']
            
            if storage == 'EMPTY_STORAGE':
                if mv_type == '641':
                    return 'BF'
                elif mv_type == '642':
                    return 'BG'
            
            if pd.notna(mv_grouping):
                key = (storage, mv_grouping)
                return storage_grouping_to_column.get(key, None)
            
            return None
        
        grouped_mb51['target_column'] = grouped_mb51.apply(get_target_column, axis=1)
        
        has_target = grouped_mb51['target_column'].notna().sum()
        no_target = grouped_mb51['target_column'].isna().sum()
        log(f"  Has target column: {has_target}/{len(grouped_mb51)}")
        log(f"  No target column: {no_target}/{len(grouped_mb51)}")
        
        # Create lookup
        log("Creating lookup dictionary...")
        mb51_lookup = {}
        
        for _, row in grouped_mb51.iterrows():
            if pd.notna(row['target_column']):
                key = (row['material'], row['plant_clean'], row['target_column'])
                if key in mb51_lookup:
                    mb51_lookup[key] += row['amount']
                else:
                    mb51_lookup[key] = row['amount']
        
        log(f"  Created lookup with {len(mb51_lookup)} keys")

        # Merge materials
        log(f"Merging materials from main file and MB51")
        
        main_file_plants = set()
        if len(existing_materials) > 0:
            main_file_plants = set(m['plant'] for m in existing_materials)
        
        all_materials = existing_materials.copy()
        existing_set = set(f"{m['material']}|{m['plant']}" for m in existing_materials)
        
        mb51_materials = df_mb51_filtered.groupby(
            ['area', 'plant_clean', 'kode_dist', 'profit_center', 'material'], 
            dropna=False
        ).size().reset_index(name='count')
        
        new_materials_added = 0
        
        for _, mb_row in mb51_materials.iterrows():
            plant_normalized = str(mb_row['plant_clean']).strip().upper()
            material = str(mb_row['material']).strip()
            key = f"{material}|{plant_normalized}"
            
            if key in existing_set:
                continue
            
            if len(main_file_plants) > 0 and plant_normalized not in main_file_plants:
                continue
            
            all_materials.append({
                'material': material, 
                'plant': plant_normalized,
                'area': mb_row['area'], 
                'kode_dist': mb_row['kode_dist'],
                'profit_center': mb_row['profit_center']
            })
            new_materials_added += 1
        
        log(f"  Existing materials: {len(existing_materials)}")
        log(f"  New materials: {new_materials_added}")
        log(f"  Total: {len(all_materials)}")

        if len(all_materials) == 0:
            raise ValueError("No materials found")

        grouped_materials = pd.DataFrame(all_materials)

        # Material descriptions
        log("Loading material descriptions...")
        material_desc_map = {}
        
        if 'Output Report INV ARUS BARANG' in sheets_dict:
            try:
                df_out = sheets_dict['Output Report INV ARUS BARANG']
                if df_out.shape[0] > 8 and df_out.shape[1] >= 7:
                    for idx in range(7, len(df_out)):
                        try:
                            row = df_out.iloc[idx]
                            mat_key = str(row.iloc[5]).strip() if len(row) > 5 else ''
                            mat_desc = str(row.iloc[6]).strip() if len(row) > 6 else ''
                            if mat_key and mat_key != 'nan' and mat_desc and mat_desc != 'nan':
                                material_desc_map[mat_key] = mat_desc
                        except:
                            continue
                    log(f"  Loaded {len(material_desc_map)} from main file")
            except Exception as e:
                log(f"  Warning: {str(e)}")
        
        if 'material_desc_mb51' in df_mb51.columns:
            desc_df = df_mb51[['material', 'material_desc_mb51']].dropna()
            desc_df = desc_df[desc_df['material_desc_mb51'].astype(str).str.strip() != '']
            desc_df = desc_df.drop_duplicates('material', keep='first')
            
            added_from_mb51 = 0
            for mat, desc in zip(desc_df['material'], desc_df['material_desc_mb51']):
                if mat not in material_desc_map:
                    material_desc_map[mat] = desc
                    added_from_mb51 += 1
            
            log(f"  Added {added_from_mb51} from MB51")
        
        log(f"  Total: {len(material_desc_map)} descriptions")

        # Create workbook
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Output_Report_INV_ARUS_BARANG_{timestamp}.xlsx"
        output_dir = os.path.join("assets", "exports")
        ensure_dir(output_dir)
        output_path = os.path.join(output_dir, filename)

        log("Creating Excel workbook...")
        wb = Workbook()
        ws = wb.active
        ws.title = "Output Report INV ARUS BARANG"
        center = Alignment(horizontal="center", vertical="center")

        # HEADER (tetap sama)
        ws["F1"], ws["F2"], ws["F3"], ws["F4"], ws["F5"], ws["F7"] = "Nama Area", "Plant", "Kode Dist", "Profit Center", "Periode", "Material"
        
        if not grouped_materials.empty:
            first_row = grouped_materials.iloc[0]
            ws["G1"], ws["G2"], ws["G3"], ws["G4"], ws["G5"] = first_row['area'], first_row['plant'], first_row['kode_dist'], first_row['profit_center'], bulan_only
        
        ws["G7"] = "Material Description"
        ws["A8"], ws["B8"], ws["C8"], ws["D8"], ws["E8"], ws["F8"] = "Nama Area", "Plant", "Kode Dist", "Profit Center", "Periode", "source data"
        
        # Row 8 labels
        ws["R8"], ws["S8"], ws["T8"], ws["U8"], ws["V8"], ws["W8"], ws["X8"], ws["Y8"] = "DTB", "BPPR", "LBP", "LBP", "DTB", "BPPR", "ALIH STATUS", "Pemusnahan"
        ws["AB8"], ws["AC8"], ws["AD8"], ws["AE8"], ws["AF8"], ws["AG8"], ws["AH8"], ws["AI8"] = "DTB", "BPPR", "LBP", "LBP", "DTB", "BPPR", "ALIH STATUS", "Pemusnahan"
        ws["AL8"], ws["AM8"], ws["AN8"], ws["AO8"], ws["AP8"], ws["AQ8"], ws["AR8"], ws["AS8"] = "DTB", "BPPR", "LBP", "LBP", "DTB", "BPPR", "ALIH STATUS", "Pemusnahan"
        ws["AV8"], ws["AW8"], ws["AX8"], ws["AY8"], ws["AZ8"], ws["BA8"], ws["BB8"], ws["BC8"] = "DTB", "BPPR", "LBP", "LBP", "DTB", "BPPR", "ALIH STATUS", "Pemusnahan"
        ws["BF8"], ws["BG8"] = "641", "642"

        ws.merge_cells("H4:M4")
        ws["H4"] = f"SALDO AWAL {bulan} {tahun}"
        ws["H4"].alignment = center
        
        ws.merge_cells("H5:J5")
        ws["H5"] = f"SALDO AWAL {prev_month} {prev_year}"
        ws["H5"].alignment = center
        
        ws.merge_cells("K5:M5")
        ws["K5"] = "SAP - MB5B"
        ws["K5"].alignment = center
        ws["N5"] = "DIFF"
        ws["N5"].alignment = center

        headers_6 = ["GS", "BS", "Grand Total", "GS", "BS", "Grand Total", "GS", "BS", "Grand Total"]
        for i, label in enumerate(headers_6, start=8):
            ws.cell(row=6, column=i, value=label).alignment = center

        for col in range(8, 17):
            ws.cell(row=7, column=col, value="S.Aw").alignment = center

        ws["R1"] = "ctrl balance MB51"
        ws.merge_cells("R5:BH5")
        ws["R5"] = "SAP - MB51"
        ws["R5"].alignment = center

        ws.merge_cells("R6:Z6")
        ws["R6"] = "GS00"
        ws["R6"].alignment = center

        gs00_movements = [
            ("R", "Terima Barang"), ("S", "Retur Beli"), ("T", "Penjualan"),
            ("U", "Retur Jual"), ("V", "Intra Gudang Masuk"), ("W", "Intra Gudang"),
            ("X", "Transfer Stock"), ("Y", "Pemusnahan"), ("Z", "Adjustment")
        ]
        for col, label7 in gs00_movements:
            ws[f"{col}7"] = label7
            ws[f"{col}7"].alignment = center

        ws.merge_cells("AB6:AJ6")
        ws["AB6"] = "BS00"
        ws["AB6"].alignment = center
        bs00_movements = [
            ("AB", "Terima Barang"), ("AC", "Retur Beli"), ("AD", "Penjualan"),
            ("AE", "Retur Jual"), ("AF", "Intra Gudang Masuk"), ("AG", "Intra Gudang"),
            ("AH", "Transfer Stock"), ("AI", "Pemusnahan"), ("AJ", "Adjustment")
        ]
        for col, label7 in bs00_movements:
            ws[f"{col}7"] = label7
            ws[f"{col}7"].alignment = center

        ws.merge_cells("AL6:AT6")
        ws["AL6"] = "AI00"
        ws["AL6"].alignment = center
        ai00_movements = [
            ("AL", "Terima Barang"), ("AM", "Retur Beli"), ("AN", "Penjualan"),
            ("AO", "Retur Jual"), ("AP", "Intra Gudang Masuk"), ("AQ", "Intra Gudang"),
            ("AR", "Transfer Stock"), ("AS", "Pemusnahan"), ("AT", "Adjustment")
        ]
        for col, label7 in ai00_movements:
            ws[f"{col}7"] = label7
            ws[f"{col}7"].alignment = center

        ws.merge_cells("AV6:BD6")
        ws["AV6"] = "TR00"
        ws["AV6"].alignment = center
        tr00_movements = [
            ("AV", "Terima Barang"), ("AW", "Retur Beli"), ("AX", "Penjualan"),
            ("AY", "Retur Jual"), ("AZ", "Intra Gudang Masuk"), ("BA", "Intra Gudang"),
            ("BB", "Transfer Stock"), ("BC", "Pemusnahan"), ("BD", "Adjustment")
        ]
        for col, label7 in tr00_movements:
            ws[f"{col}7"] = label7
            ws[f"{col}7"].alignment = center

        ws.merge_cells("BF6:BH6")
        ws["BF6"] = "641 dan 642 tanpa sloc"
        ws["BF6"].alignment = center
        ws["BF7"], ws["BG7"], ws["BH7"] = "Intra Gudang", "Intra Gudang", "CEK"
        ws["BI3"], ws["BI4"] = "-->stock in transit", "jika selisih cek ke MB5T"

        # END STOCK
        ws.merge_cells("BK4:BP4")
        ws["BK4"] = f"END STOCK {prev_month} {prev_year}"
        ws["BK4"].alignment = center
        ws.merge_cells("BK5:BM5")
        ws["BK5"] = "SALDO AKHIR"
        ws["BK5"].alignment = center
        ws.merge_cells("BN5:BP5")
        ws["BN5"] = "SAP - MB5B"
        ws["BN5"].alignment = center
        ws["BQ5"] = "DIFF"
        ws["BQ5"].alignment = center

        ws["BK6"], ws["BL6"], ws["BM6"] = "GS00", "BS00", "Grand Total"
        ws["BN6"], ws["BO6"], ws["BP6"] = "GS", "BS", "Grand Total"
        ws["BQ6"], ws["BR6"], ws["BS6"] = "GS", "BS", "Grand Total"

        for col in range(63, 72):
            ws.cell(row=6, column=col).alignment = center
            ws.cell(row=7, column=col, value="S.Ak").alignment = center

        ws["BT7"] = "CEK SELISIH VS BULAN LALU"
        ws["BU7"] = "kalo ada selisih atas inputan LOG1, LOG2 -> konfirmasi pa Reza utk diselesaikan"

        ws.merge_cells("BV5:BX5")
        ws["BV5"] = "STOCK - EDS"
        ws["BV5"].alignment = center
        ws["BY5"] = "DIFF"
        ws["BY5"].alignment = center

        ws["BV6"], ws["BW6"], ws["BX6"] = "GS", "BS", "Grand Total"
        ws["BY6"], ws["BZ6"], ws["CA6"] = "GS", "BS", "Grand Total"

        for col in range(74, 80):
            ws.cell(row=6, column=col).alignment = center
            ws.cell(row=7, column=col, value="S.Ak").alignment = center

        # TAMBAHAN: BASO HEADERS (kolom CC, CD, CE)
        ws.merge_cells("CC5:CE5")
        ws["CC5"] = "STOCK - BASO"
        ws["CC5"].alignment = center

        ws["CC6"], ws["CD6"], ws["CE6"] = "GS", "BS", "Grand Total"
        for col in range(get_column_index("CC"), get_column_index("CE") + 1):
            ws.cell(row=6, column=col).alignment = center
            ws.cell(row=7, column=col, value="S.Ak").alignment = center

        # BODY CALCULATION
        log("Calculating body rows...")
        write_row = 9
        totals = defaultdict(float)
        
        # Initialize tracking
        eds_hits = {'GS': 0, 'BS': 0, 'total_queries': 0}
        baso_hits = {'GS': 0, 'BS': 0, 'total_queries': 0}
        
        num_materials = len(grouped_materials)
        all_target_columns = ["R", "S", "T", "U", "V", "W", "X", "Y", "Z",
                              "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ",
                              "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT",
                              "AV", "AW", "AX", "AY", "AZ", "BA", "BB", "BC", "BD",
                              "BF", "BG"]
        
        for idx in range(num_materials):
            if idx % 100 == 0:
                log(f"  Processing {idx}/{num_materials}")
            
            mat_row = grouped_materials.iloc[idx]
            area = str(mat_row['area'])
            plant = str(mat_row['plant']).strip().upper()
            kode_dist = str(mat_row['kode_dist'])
            profit_center = str(mat_row['profit_center'])
            material = str(mat_row['material'])
            material_desc = material_desc_map.get(material, "")
            
            plant_exists_in_mb51 = plant in mb51_plants

            # Write basic info
            ws.cell(row=write_row, column=1, value=area)
            ws.cell(row=write_row, column=2, value=plant)
            ws.cell(row=write_row, column=3, value=kode_dist)
            ws.cell(row=write_row, column=4, value=profit_center)
            ws.cell(row=write_row, column=5, value=bulan_only)
            ws.cell(row=write_row, column=6, value=material)
            ws.cell(row=write_row, column=7, value=material_desc)

            # SALDO AWAL
            h9 = sheet_cache.get_saldo_awal(material, plant, "GS")
            i9 = sheet_cache.get_saldo_awal(material, plant, "BS")
            j9_formula = f"=H{write_row}+I{write_row}"
            
            ws.cell(row=write_row, column=8, value=h9)
            ws.cell(row=write_row, column=9, value=i9)
            ws.cell(row=write_row, column=10, value=j9_formula)

            k9 = sheet_cache.get_mb5b_awal(material, plant, "GS")
            l9 = sheet_cache.get_mb5b_awal(material, plant, "BS")
            m9_formula = f"=SUM(K{write_row}:L{write_row})"
            
            ws.cell(row=write_row, column=11, value=k9)
            ws.cell(row=write_row, column=12, value=l9)
            ws.cell(row=write_row, column=13, value=m9_formula)

            n9_formula = f"=H{write_row}-K{write_row}"
            o9_formula = f"=I{write_row}-L{write_row}"
            p9_formula = f"=N{write_row}+O{write_row}"
            
            ws.cell(row=write_row, column=14, value=n9_formula)
            ws.cell(row=write_row, column=15, value=o9_formula)
            ws.cell(row=write_row, column=16, value=p9_formula)

            # Write MB51 columns
            if plant_exists_in_mb51:
                for target_col in all_target_columns:
                    lookup_key = (material, plant, target_col)
                    amount = mb51_lookup.get(lookup_key, 0.0)
                    ws.cell(row=write_row, column=get_column_index(target_col), value=amount)
                    totals[target_col] += amount
            else:
                for target_col in all_target_columns:
                    ws.cell(row=write_row, column=get_column_index(target_col), value=0)

            # BH formula
            bh9_formula = f"=V{write_row}-BF{write_row}-BG{write_row}"
            ws.cell(row=write_row, column=get_column_index("BH"), value=bh9_formula)

            # END STOCK
            bk9_formula = f"=H{write_row}+SUM(R{write_row}:Z{write_row})+SUM(AL{write_row}:BD{write_row})"
            bl9_formula = f"=I{write_row}+SUM(AB{write_row}:AJ{write_row})"
            bm9_formula = f"=BK{write_row}+BL{write_row}"
            
            ws.cell(row=write_row, column=get_column_index("BK"), value=bk9_formula)
            ws.cell(row=write_row, column=get_column_index("BL"), value=bl9_formula)
            ws.cell(row=write_row, column=get_column_index("BM"), value=bm9_formula)

            # SAP - MB5B
            bn9 = sheet_cache.get_mb5b(material, plant, "GS")
            bo9 = sheet_cache.get_mb5b(material, plant, "BS")
            bp9_formula = f"=SUM(BN{write_row}:BO{write_row})"
            
            ws.cell(row=write_row, column=get_column_index("BN"), value=bn9)
            ws.cell(row=write_row, column=get_column_index("BO"), value=bo9)
            ws.cell(row=write_row, column=get_column_index("BP"), value=bp9_formula)

            # DIFF
            bq9_formula = f"=BK{write_row}-BN{write_row}"
            br9_formula = f"=BL{write_row}-BO{write_row}"
            bs9_formula = f"=BQ{write_row}+BR{write_row}"
            
            ws.cell(row=write_row, column=get_column_index("BQ"), value=bq9_formula)
            ws.cell(row=write_row, column=get_column_index("BR"), value=br9_formula)
            ws.cell(row=write_row, column=get_column_index("BS"), value=bs9_formula)

            bt9_formula = f"=P{write_row}-BS{write_row}"
            ws.cell(row=write_row, column=get_column_index("BT"), value=bt9_formula)

            # STOCK - EDS
            bv9 = sheet_cache.get_eds(material, plant, "GS")
            bw9 = sheet_cache.get_eds(material, plant, "BS")
            bx9_formula = f"=BV{write_row}+BW{write_row}"
            
            eds_hits['total_queries'] += 2
            if bv9 != 0:
                eds_hits['GS'] += 1
            if bw9 != 0:
                eds_hits['BS'] += 1
            
            ws.cell(row=write_row, column=get_column_index("BV"), value=bv9)
            ws.cell(row=write_row, column=get_column_index("BW"), value=bw9)
            ws.cell(row=write_row, column=get_column_index("BX"), value=bx9_formula)

            by9_formula = f"=BN{write_row}-BV{write_row}"
            bz9_formula = f"=BO{write_row}-BW{write_row}"
            ca9_formula = f"=BY{write_row}+BZ{write_row}"
            
            ws.cell(row=write_row, column=get_column_index("BY"), value=by9_formula)
            ws.cell(row=write_row, column=get_column_index("BZ"), value=bz9_formula)
            ws.cell(row=write_row, column=get_column_index("CA"), value=ca9_formula)

            # TAMBAHAN: STOCK - BASO (kolom CC, CD, CE)
            cc9 = sheet_cache.get_baso(material, plant, "GS")
            cd9 = sheet_cache.get_baso(material, plant, "BS")
            ce9_formula = f"=CC{write_row}+CD{write_row}"
            
            baso_hits['total_queries'] += 2
            if cc9 != 0:
                baso_hits['GS'] += 1
            if cd9 != 0:
                baso_hits['BS'] += 1
            
            ws.cell(row=write_row, column=get_column_index("CC"), value=cc9)
            ws.cell(row=write_row, column=get_column_index("CD"), value=cd9)
            ws.cell(row=write_row, column=get_column_index("CE"), value=ce9_formula)

            write_row += 1

        log(f"Total rows written: {write_row - 9}")
        
        # Log EDS usage
        log(f"  === EDS Cache Usage ===")
        log(f"  Total queries: {eds_hits['total_queries']}")
        log(f"  GS hits: {eds_hits['GS']}, BS hits: {eds_hits['BS']}")
        
        # Log BASO usage
        log(f"  === BASO Cache Usage ===")
        log(f"  Total queries: {baso_hits['total_queries']}")
        log(f"  GS hits: {baso_hits['GS']}, BS hits: {baso_hits['BS']}")

        # Write formulas
        log("Writing formulas...")
        last_row = write_row - 1
        
        sum_columns = ["R", "S", "T", "U", "V", "W", "X", "Y", "Z",
                      "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ",
                      "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS", "AT",
                      "AV", "AW", "AX", "AY", "AZ", "BA", "BB", "BC", "BD",
                      "BF", "BG", "BH"]
        
        for col in sum_columns:
            ws[f"{col}3"] = f"=SUM({col}9:{col}{last_row})"
        
        # S1 - ctrl balance
        if len(main_file_plants) > 0:
            mb51_for_s1 = df_mb51_filtered[df_mb51_filtered['plant_clean'].isin(main_file_plants)]
            mb51_total_amount = mb51_for_s1['amount'].sum()
        else:
            mb51_total_amount = df_mb51_filtered['amount'].sum()
        
        sum_r3_bf3 = sum([totals.get(col, 0) for col in sum_columns if col != "BH"])
        s1_value = round(mb51_total_amount - sum_r3_bf3, 2)
        ws["S1"] = s1_value
        log(f"  S1 = {s1_value:.2f}")
        
        # BB2 formula
        ws["BB2"] = "=X3+AH3+AR3+BB3"
        
        # BP2 calculation
        sum_bp = 0.0
        for row in range(9, write_row):
            cell_val = ws.cell(row=row, column=get_column_index("BP")).value
            if isinstance(cell_val, (int, float)):
                sum_bp += cell_val
        
        sum_mb5b_pq = 0.0
        if '13. MB5B' in sheets_dict:
            df_mb5b_sheet = sheets_dict['13. MB5B']
            try:
                if df_mb5b_sheet.shape[1] > 16:
                    sum_p = pd.to_numeric(df_mb5b_sheet.iloc[:, 15], errors='coerce').fillna(0).sum()
                    sum_q = pd.to_numeric(df_mb5b_sheet.iloc[:, 16], errors='coerce').fillna(0).sum()
                    sum_mb5b_pq = sum_p + sum_q
            except Exception as e:
                log(f"  Warning: {str(e)}")
        
        bp2_value = round(sum_bp - sum_mb5b_pq, 2)
        ws["BP2"] = bp2_value
        log(f"  BP2 = {bp2_value:.2f}")

        # Formatting
        log("Formatting...")
        for i in range(1, 85):  # Extended untuk BASO columns
            ws.column_dimensions[get_column_letter(i)].width = 12
        
        ws.column_dimensions['Q'].width = 2
        ws.column_dimensions['AA'].width = 2
        ws.column_dimensions['AK'].width = 2
        ws.column_dimensions['AU'].width = 2
        ws.column_dimensions['BE'].width = 2
        ws.column_dimensions['BJ'].width = 4
        ws.column_dimensions['CB'].width = 2  # Space before BASO

        ws.freeze_panes = "H9"

        for row in range(9, write_row):
            for col in range(8, 85):  # Extended untuk BASO
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
        
        for row in [2, 3]:
            for col in range(18, 85):  # Extended untuk BASO
                ws.cell(row=row, column=col).number_format = '#,##0'

        # Save
        log(f"Saving workbook...")
        wb.save(output_path)
        
        if not os.path.exists(output_path):
            raise Exception(f"File was not created")
        
        file_size = os.path.getsize(output_path)
        log(f"✓ File created: {file_size:,} bytes")
        
        result = {
            "success": True,
            "output_path": output_path,
            "rows_written": write_row - 9,
            "report_month": f"{bulan} {tahun}",
            "total_materials": len(grouped_materials),
            "file_size": file_size,
            "timestamp": timestamp,
            "unmapped_count": int(unmapped_count),
            "no_target_count": int(no_target),
            "eds_hits": {
                "total_queries": eds_hits['total_queries'],
                "gs_hits": eds_hits['GS'],
                "bs_hits": eds_hits['BS']
            },
            "baso_hits": {
                "total_queries": baso_hits['total_queries'],
                "gs_hits": baso_hits['GS'],
                "bs_hits": baso_hits['BS']
            },
            "baso_available": baso_path is not None
        }
        
        print(json.dumps(result))
        sys.stdout.flush()
        log("✓ Report completed successfully with BASO support!")

    except Exception as e:
        tb = traceback.format_exc()
        log(f"ERROR: {str(e)}")
        log(f"Traceback:\n{tb}")
        
        error_result = {
            "success": False,
            "error": str(e),
            "trace": tb
        }
        
        print(json.dumps(error_result))
        sys.stdout.flush()
        sys.exit(1)


if __name__ == "__main__":
    main()