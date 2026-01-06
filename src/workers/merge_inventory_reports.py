# merge_inventory_reports.py - NETWORK TIMEOUT RESISTANT
# Major Features:
# 1. Chunked processing with progress updates
# 2. Keeps network connection alive with heartbeats
# 3. Handles 275+ files without timeout
# 4. Recovery from partial failures

import sys
import json
import os
import datetime
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook

def log(msg):
    """Log to stderr"""
    print(f"[merge-worker] {msg}", file=sys.stderr, flush=True)

def send_progress(stage, current, total, message=""):
    """Send progress update to keep connection alive"""
    progress_data = {
        "type": "progress",
        "stage": stage,
        "current": current,
        "total": total,
        "percentage": round((current / total * 100), 1) if total > 0 else 0,
        "message": message
    }
    print(json.dumps(progress_data), flush=True)

def ensure_dir(path):
    os.makedirs(path, exist_ok=True)

def copy_cell_style(source_cell, target_cell):
    """Copy styling from source to target cell"""
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.alignment = source_cell.alignment.copy()

def read_file_data(file_path, file_idx, total_files):
    """Read data from a single file - optimized"""
    try:
        wb = load_workbook(file_path, data_only=True, keep_links=False)
        
        sheet_name = "Output Report INV ARUS BARANG"
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]
        
        # Extract metadata
        plant_code = None
        try:
            plant_cell = ws.cell(row=2, column=7).value
            if plant_cell and str(plant_cell).strip() and str(plant_cell).strip() != 'nan':
                plant_code = str(plant_cell).strip()
        except:
            pass
        
        # Extract S1 and BL2
        s1_value = 0.0
        bl2_value = 0.0
        try:
            s1_cell = ws.cell(row=1, column=19).value
            if s1_cell and isinstance(s1_cell, (int, float)):
                s1_value = float(s1_cell)
        except:
            pass
        
        try:
            bl2_cell = ws.cell(row=2, column=64).value
            if bl2_cell and isinstance(bl2_cell, (int, float)):
                bl2_value = float(bl2_cell)
        except:
            pass
        
        # Read up to column BZ (78 columns)
        max_col = min(ws.max_column, 78)
        total_rows = ws.max_row
        
        # Read data rows
        data_rows = []
        for row_idx in range(9, total_rows + 1):
            material_val = ws.cell(row=row_idx, column=6).value
            if not material_val or str(material_val).strip() == '' or str(material_val).strip() == 'nan':
                continue
            
            row_data = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
            data_rows.append(row_data)
        
        wb.close()
        
        return {
            'file_path': file_path,
            'file_idx': file_idx,
            'plant_code': plant_code,
            'data_rows': data_rows,
            'max_col': max_col,
            's1_value': s1_value,
            'bl2_value': bl2_value,
            'filename': os.path.basename(file_path)
        }
        
    except Exception as e:
        log(f"ERROR reading file {file_idx + 1}: {str(e)}")
        return {
            'file_path': file_path,
            'file_idx': file_idx,
            'error': str(e),
            'filename': os.path.basename(file_path)
        }

def write_batch_rows(ws_output, start_row, data_rows, max_col):
    """Write multiple rows in batch"""
    current_row = start_row
    
    for row_data in data_rows:
        for col_idx in range(1, min(len(row_data) + 1, max_col + 1)):
            cell = ws_output.cell(row=current_row, column=col_idx)
            cell.value = row_data[col_idx - 1]
            
            if col_idx >= 8 and isinstance(row_data[col_idx - 1], (int, float)):
                cell.number_format = '#,##0'
        
        current_row += 1
    
    return current_row

def main():
    try:
        payload = json.load(sys.stdin)
        file_paths = payload.get("file_paths", [])
        
        if not file_paths or len(file_paths) == 0:
            raise ValueError("No file paths provided")
        
        total_files = len(file_paths)
        log(f"Starting merge for {total_files} files")
        send_progress("init", 0, total_files, f"Initializing merge for {total_files} files")
        
        # Validate files
        for fp in file_paths:
            if not os.path.exists(fp):
                raise FileNotFoundError(f"File not found: {fp}")
        
        # STAGE 1: Read files in parallel with chunked progress
        log("STAGE 1: Reading files...")
        send_progress("reading", 0, total_files, "Starting file reading")
        
        max_workers = min(8, total_files, os.cpu_count() or 4)
        file_data_list = []
        completed_count = 0
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(read_file_data, fp, idx, total_files): idx
                for idx, fp in enumerate(file_paths)
            }
            
            for future in as_completed(futures):
                file_data = future.result()
                completed_count += 1
                
                # Send progress every file (keeps connection alive)
                filename = file_data.get('filename', 'unknown')
                if 'error' in file_data:
                    log(f"WARNING: Skipped {filename}")
                    send_progress("reading", completed_count, total_files, f"Skipped {filename} (error)")
                else:
                    rows = len(file_data.get('data_rows', []))
                    file_data_list.append(file_data)
                    send_progress("reading", completed_count, total_files, f"Read {filename} ({rows} rows)")
        
        file_data_list.sort(key=lambda x: x['file_idx'])
        
        if not file_data_list:
            raise ValueError("No valid files to merge")
        
        log(f"Successfully read {len(file_data_list)} files")
        
        # Aggregate metadata
        plant_codes = set()
        total_s1 = 0.0
        total_bl2 = 0.0
        total_data_rows = 0
        
        for fd in file_data_list:
            if fd.get('plant_code'):
                plant_codes.add(fd['plant_code'])
            total_s1 += fd.get('s1_value', 0.0)
            total_bl2 += fd.get('bl2_value', 0.0)
            total_data_rows += len(fd['data_rows'])
        
        log(f"Total rows: {total_data_rows} | Plants: {len(plant_codes)} | S1: {total_s1:.2f} | BL2: {total_bl2:.2f}")
        send_progress("aggregation", len(file_data_list), total_files, f"Total: {total_data_rows} rows from {len(plant_codes)} plants")
        
        if total_data_rows == 0:
            raise ValueError("No data rows found")
        
        # STAGE 2: Create workbook and copy header
        log("STAGE 2: Creating workbook...")
        send_progress("creating", 0, 3, "Creating output workbook")
        
        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = "Output Report INV ARUS BARANG"
        
        send_progress("creating", 1, 3, "Copying header")
        
        # Copy header
        first_file = file_paths[0]
        wb_first = load_workbook(first_file, data_only=False)
        sheet_name = "Output Report INV ARUS BARANG"
        ws_first = wb_first[sheet_name] if sheet_name in wb_first.sheetnames else wb_first.worksheets[0]
        
        max_col = ws_first.max_column
        for row_idx in range(1, 9):
            for col_idx in range(1, max_col + 1):
                source_cell = ws_first.cell(row=row_idx, column=col_idx)
                target_cell = ws_output.cell(row=row_idx, column=col_idx)
                target_cell.value = source_cell.value
                copy_cell_style(source_cell, target_cell)
        
        # Copy column widths and merged cells
        for col_letter in ws_first.column_dimensions:
            ws_output.column_dimensions[col_letter].width = ws_first.column_dimensions[col_letter].width
        
        for merged_range in ws_first.merged_cells.ranges:
            if merged_range.min_row <= 8:
                ws_output.merge_cells(str(merged_range))
        
        wb_first.close()
        
        # Update header metadata
        ws_output["G1"].value = "Merge Report"
        ws_output["G2"].value = ", ".join(sorted(plant_codes)) if plant_codes else "-"
        ws_output["G3"].value = "-"
        ws_output["G4"].value = "-"
        
        send_progress("creating", 2, 3, "Header copied")
        
        # STAGE 3: Write data with chunked progress
        log("STAGE 3: Writing data rows...")
        current_row = 9
        rows_written = 0
        
        for idx, file_data in enumerate(file_data_list):
            data_rows = file_data['data_rows']
            max_col = file_data['max_col']
            filename = file_data['filename']
            
            # Write rows
            current_row = write_batch_rows(ws_output, current_row, data_rows, max_col)
            rows_written += len(data_rows)
            
            # Send progress every file
            send_progress("writing", idx + 1, len(file_data_list), 
                         f"Written {filename} ({rows_written}/{total_data_rows} rows)")
        
        last_data_row = current_row - 1
        total_rows_written = last_data_row - 8
        
        log(f"Written {total_rows_written} rows (row 9 to {last_data_row})")
        
        # Apply freeze panes
        ws_output.freeze_panes = "H9"
        
        # STAGE 4: Update formulas
        log("STAGE 4: Updating formulas...")
        send_progress("formulas", 0, 1, "Updating formulas")
        
        sum_columns = ["R", "S", "T", "U", "V", "W", "X", "Y", "Z",
                      "AB", "AC", "AD", "AE", "AF", "AG", "AH",
                      "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ",
                      "AS", "AT", "AU", "AV", "AW", "AX", "AY", "AZ",
                      "BB", "BC", "BD"]
        
        for col_letter in sum_columns:
            ws_output[f"{col_letter}3"].value = f"=SUM({col_letter}9:{col_letter}{last_data_row})"
            ws_output[f"{col_letter}3"].number_format = '#,##0'
        
        ws_output["S1"].value = total_s1
        ws_output["S1"].number_format = '#,##0'
        
        ws_output["AX2"].value = "=X3+AF3+AO3+AX3"
        ws_output["AX2"].number_format = '#,##0'
        
        ws_output["BL2"].value = total_bl2
        ws_output["BL2"].number_format = '#,##0'
        
        send_progress("formulas", 1, 1, "Formulas updated")
        
        # STAGE 5: Save file
        log("STAGE 5: Saving file...")
        send_progress("saving", 0, 1, "Saving consolidated file")
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Consolidated_Report_INV_ARUS_BARANG_{timestamp}.xlsx"
        output_dir = os.path.join("assets", "exports")
        ensure_dir(output_dir)
        output_path = os.path.join(output_dir, filename)
        
        wb_output.save(output_path)
        wb_output.close()
        
        if not os.path.exists(output_path):
            raise Exception(f"File was not created")
        
        file_size = os.path.getsize(output_path)
        log(f"SUCCESS - Size: {file_size:,} bytes")
        
        send_progress("complete", 1, 1, "Merge completed successfully")
        
        # Send final result
        result = {
            "success": True,
            "output_path": output_path,
            "total_files_merged": len(file_data_list),
            "total_data_rows": total_rows_written,
            "plant_codes": sorted(list(plant_codes)),
            "file_size": file_size,
            "timestamp": timestamp
        }
        
        print(json.dumps(result))
        sys.stdout.flush()
        
    except Exception as e:
        tb = traceback.format_exc()
        log(f"ERROR: {str(e)}")
        
        error_result = {
            "success": False,
            "error": str(e),
            "trace": tb
        }
        
        print(json.dumps(error_result))
        sys.stdout.flush()
        sys.exit(1)


if __name__ == "__main__":
    main()